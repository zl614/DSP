#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SpeedX 包裹清单拆分器 + 按车队发邮件 —— 单文件本地程序
=========================================================
双击 exe 后：① 起一个本地服务 ② 自动打开浏览器。

做两件事：
  1) 把一份 Excel 按「Last delivery fleet」拆成每个车队一个文件（可打包下载 ZIP）
  2) 把各自那份直接发给对应车队 —— 走 Microsoft Graph，用你自己的 Outlook 账号

为什么发邮件非要经过这个程序：
  新版 Outlook 砍掉了 COM 接口，本地脚本（.bat / PowerShell / AppleScript）驱动不了它，
  .eml 它也不认；浏览器又不能直接连 SMTP，mailto 也不能带附件。
  能发信的只剩 Microsoft Graph API —— 它要跑 OAuth，就需要一个后端，就是这个程序。

不存密码：走 OAuth（设备码流程），登录态只在内存里，关掉程序就没了。

本地跑（可选，需要 Python）：
    pip install flask requests
    python speedx_relay.py
打 Windows exe：推到 GitHub，Actions 里的 build-windows-exe 会自动打包。
"""
import os
import sys
import json
import time
import re
import base64
import hashlib
import secrets
import socket
import threading
import webbrowser
from urllib.parse import urlencode

from flask import Flask, request, Response, send_file
import requests

# ===================== 配置 =====================
CFG = {
    # 端口必须固定：Azure 里登记的重定向 URI 是写死端口的，浮动端口会导致登录失败。
    "port": 5058,
}
REDIRECT_URI = f"http://localhost:{CFG['port']}/mail/callback"   # ← Azure 里要登记这一条

# ===================== Outlook 邮件（Microsoft Graph）=====================
# OAuth 和发信全部用 requests 手写，不引入新依赖。
#
# 配置方式（优先级从高到低）：
#   1) 环境变量 SPEEDX_GRAPH_CLIENT_ID / SPEEDX_GRAPH_TENANT_ID
#   2) 与本程序（或 exe）同目录的 graph_config.json
#   3) 下面 GRAPH_CFG 里的默认值
# 没配 client_id 也不影响：拆分、下载 ZIP 照常用，只是「发送」按钮是灰的。
GRAPH_CFG = {
    "client_id": "",
    "tenant_id": "common",
    "allow_draft": True,   # True=可建草稿(需 Mail.ReadWrite)；False=只发送(仅需 Mail.Send)
}
GRAPH_LOGIN = "https://login.microsoftonline.com"
GRAPH_API = "https://graph.microsoft.com/v1.0"

GRAPH_TOKEN = {"access": None, "refresh": None, "exp": 0.0, "user": ""}
GRAPH_FLOW = {"device_code": None, "interval": 5, "expires": 0.0}   # 设备码（备用登录方式）
OAUTH = {"verifier": None, "state": None}                            # 授权码 + PKCE（主登录方式）

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

PAGE = "fleet_splitter.html"


def _app_dir():
    """exe 时取 exe 所在目录（改配置/改页面都不用重新打包）；开发时取脚本目录。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(rel):
    """打包后从 _MEIPASS 读，开发时从脚本目录读。"""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def find_asset(rel):
    """找页面文件：**exe 同目录优先**，其次才是打包进 exe 的那份。
    这样改页面不用重新打包 —— 把新的 html 丢到 exe 旁边、刷新浏览器就行。"""
    side = os.path.join(_app_dir(), rel)
    if os.path.exists(side):
        return side
    return resource_path(rel)


def asset_origin(rel):
    if os.path.exists(os.path.join(_app_dir(), rel)):
        return "exe 同目录"
    return "exe 内置" if os.path.exists(resource_path(rel)) else "缺失 !"


def _graph_scopes():
    s = ["https://graph.microsoft.com/Mail.Send",
         "https://graph.microsoft.com/User.Read",
         "offline_access"]
    if GRAPH_CFG["allow_draft"]:
        s.insert(1, "https://graph.microsoft.com/Mail.ReadWrite")
    return " ".join(s)


def _authority():
    return f"{GRAPH_LOGIN}/{(GRAPH_CFG['tenant_id'] or 'common').strip()}"


def load_graph_cfg():
    cid = os.environ.get("SPEEDX_GRAPH_CLIENT_ID")
    tid = os.environ.get("SPEEDX_GRAPH_TENANT_ID")
    if cid:
        GRAPH_CFG["client_id"] = cid.strip()
    if tid:
        GRAPH_CFG["tenant_id"] = tid.strip()
    p = os.path.join(_app_dir(), "graph_config.json")
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                j = json.load(f)
            if j.get("client_id"):
                GRAPH_CFG["client_id"] = str(j["client_id"]).strip()
            if j.get("tenant_id"):
                GRAPH_CFG["tenant_id"] = str(j["tenant_id"]).strip()
            if "allow_draft" in j:
                GRAPH_CFG["allow_draft"] = bool(j["allow_draft"])
        except Exception as e:
            print(f"[mail] graph_config.json 读取失败：{e}")
    return bool(GRAPH_CFG["client_id"])


def _store_token(j):
    GRAPH_TOKEN["access"] = j.get("access_token")
    GRAPH_TOKEN["refresh"] = j.get("refresh_token") or GRAPH_TOKEN["refresh"]
    GRAPH_TOKEN["exp"] = time.time() + int(j.get("expires_in", 3600)) - 120
    try:
        me = requests.get(f"{GRAPH_API}/me",
                          headers={"Authorization": "Bearer " + GRAPH_TOKEN["access"]},
                          timeout=20).json()
        GRAPH_TOKEN["user"] = me.get("mail") or me.get("userPrincipalName") or ""
    except Exception:
        pass


def graph_token():
    """拿可用的 access token；过期就用 refresh token 续，不用你重新登录。"""
    if GRAPH_TOKEN["access"] and time.time() < GRAPH_TOKEN["exp"]:
        return GRAPH_TOKEN["access"]
    if GRAPH_TOKEN["refresh"]:
        r = requests.post(f"{_authority()}/oauth2/v2.0/token", data={
            "grant_type": "refresh_token",
            "client_id": GRAPH_CFG["client_id"],
            "refresh_token": GRAPH_TOKEN["refresh"],
            "scope": _graph_scopes(),
        }, timeout=30)
        j = r.json()
        if "access_token" in j:
            _store_token(j)
            return GRAPH_TOKEN["access"]
    raise RuntimeError("邮箱未登录或登录已过期，请重新登录。")


def _recips(s):
    return [{"emailAddress": {"address": a.strip()}}
            for a in re.split(r"[;,]", s or "") if a.strip()]


# ===================== Web =====================
app = Flask(__name__)


@app.after_request
def cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return resp


import io, copy, hashlib, zipfile, threading
from collections import OrderedDict

# 解析一次原始 xlsx 很慢（大文件带样式约 10-15s），所以按内容哈希缓存，
# 同一份文件后续拆分（下载 ZIP、逐封发邮件）直接复用，不重复解析。
_WB_CACHE = {"key": None, "sws": None, "header_cells": None, "col_widths": None,
             "row_heights": None, "max_col": None, "groups": None, "sheet": None}
_WB_LOCK = threading.Lock()


def _sanitize(name):
    out = "".join(c if c not in '\\/:*?"<>|' else "_" for c in str(name)).strip()
    out = out.rstrip(". ")
    return (out or "unnamed")[:120]


def _load_source(src_bytes, sheet_name, col_idx, header_row):
    """解析原始工作簿 → 缓存表头样式、列宽、按车队分好的行号。返回缓存 key。"""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    key = hashlib.sha1(src_bytes).hexdigest() + f"|{sheet_name}|{col_idx}|{header_row}"
    with _WB_LOCK:
        if _WB_CACHE["key"] == key:
            return key
        wb = load_workbook(io.BytesIO(src_bytes))
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        sws = wb[sheet_name]
        max_col = sws.max_column
        col = col_idx + 1
        header_cells = [[sws.cell(r, i) for i in range(1, max_col + 1)]
                        for r in range(1, header_row + 1)]
        col_widths = {}
        row_heights = {}
        # 复制所有有记录的列宽（含 H 之后的空列，比如 I），不只到 max_col
        for L, dim in sws.column_dimensions.items():
            if dim.width:
                col_widths[L] = dim.width
        for r in range(1, header_row + 1):
            if r in sws.row_dimensions and sws.row_dimensions[r].height:
                row_heights[r] = sws.row_dimensions[r].height
        groups = OrderedDict()
        for r in range(header_row + 1, sws.max_row + 1):
            v = sws.cell(r, col).value
            k = "" if v is None else str(v).strip()
            if k == "":
                continue
            groups.setdefault(k, []).append(r)
        _WB_CACHE.update(key=key, sws=sws, header_cells=header_cells, col_widths=col_widths,
                         row_heights=row_heights, max_col=max_col, groups=groups, sheet=sheet_name)
        return key


def _copy_cell(sc, dc):
    dc.value = sc.value
    if sc.has_style:
        dc.font = copy.copy(sc.font)
        dc.fill = copy.copy(sc.fill)
        dc.border = copy.copy(sc.border)
        dc.alignment = copy.copy(sc.alignment)
        dc.number_format = sc.number_format
        dc.protection = copy.copy(sc.protection)


def _build_fleet(fleet_value):
    """从缓存里生成某个车队的 xlsx（保留原格式）。返回 bytes。"""
    from openpyxl import Workbook
    C = _WB_CACHE
    sws, max_col = C["sws"], C["max_col"]
    rownums = C["groups"].get(fleet_value, [])
    dwb = Workbook()
    dws = dwb.active
    dws.title = (C["sheet"] or "Sheet1")[:31]
    for L, w in C["col_widths"].items():
        dws.column_dimensions[L].width = w
    dr = 1
    for hr_i, hc in enumerate(C["header_cells"], start=1):
        for i, sc in enumerate(hc, start=1):
            _copy_cell(sc, dws.cell(dr, i))
        if hr_i in C["row_heights"]:
            dws.row_dimensions[dr].height = C["row_heights"][hr_i]
        dr += 1
    for rn in rownums:
        for i in range(1, max_col + 1):
            _copy_cell(sws.cell(rn, i), dws.cell(dr, i))
        dr += 1
    b = io.BytesIO()
    dwb.save(b)
    return b.getvalue()


@app.route("/split-all", methods=["POST"])
def split_all():
    """收原始 xlsx，一次性返回「每个车队一个文件（带原格式）+ 汇总」的 ZIP。"""
    f = request.files.get("file")
    if not f:
        return {"ok": False, "msg": "没收到文件"}, 400
    try:
        sheet = request.form["sheet"]
        col_idx = int(request.form["colIdx"])
        header_row = int(request.form.get("headerRow", "1"))
        base = request.form.get("base", "split")
        mapping = json.loads(request.form.get("emails", "{}"))   # {fleet: email} 仅用于汇总表
        src = f.read()
        _load_source(src, sheet, col_idx, header_row)
    except Exception as e:
        return {"ok": False, "msg": f"{type(e).__name__}: {e}"}, 500

    C = _WB_CACHE
    used = {}
    zbuf = io.BytesIO()
    summary = [["File", "Fleet", "Rows", "Email"]]
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as z:
        for fleet, rns in C["groups"].items():
            fn = _sanitize(fleet)
            low = fn.lower()
            if low in used:
                used[low] += 1
                fn = f"{fn}_{used[low]}"
            else:
                used[low] = 1
            z.writestr(fn + ".xlsx", _build_fleet(fleet))
            summary.append([fn + ".xlsx", fleet, len(rns), mapping.get(fleet, "")])
        # 汇总表（用 openpyxl 简单写，无需格式）
        from openpyxl import Workbook
        swb = Workbook()
        sws2 = swb.active
        sws2.title = "Summary"
        for row in summary:
            sws2.append(row)
        sb = io.BytesIO()
        swb.save(sb)
        z.writestr("_SUMMARY.xlsx", sb.getvalue())
    zbuf.seek(0)
    return send_file(zbuf, mimetype="application/zip", as_attachment=True,
                     download_name=f"{_sanitize(base)}__by_fleet.zip")


@app.route("/split-one", methods=["POST"])
def split_one():
    """返回单个车队的 xlsx（带原格式）——发邮件时一封一封取，复用缓存所以很快。"""
    f = request.files.get("file")
    try:
        sheet = request.form["sheet"]
        col_idx = int(request.form["colIdx"])
        header_row = int(request.form.get("headerRow", "1"))
        fleet = request.form["fleet"]
        fname = request.form.get("fileName", "fleet.xlsx")
        # 有文件就（重新）建缓存；没有就用现有缓存（前端发邮件时不必每封都重传大文件）
        if f is not None:
            _load_source(f.read(), sheet, col_idx, header_row)
        elif _WB_CACHE["key"] is None:
            return {"ok": False, "msg": "缓存已失效，请重新选文件"}, 400
        data = _build_fleet(fleet)
    except Exception as e:
        return {"ok": False, "msg": f"{type(e).__name__}: {e}"}, 500
    return send_file(io.BytesIO(data), mimetype=XLSX_MIME, as_attachment=True, download_name=fname)


@app.route("/")
@app.route("/split")            # 老链接也留着
def index():
    try:
        with open(find_asset(PAGE), encoding="utf-8") as f:
            return Response(f.read(), mimetype="text/html; charset=utf-8")
    except FileNotFoundError:
        return Response(f"找不到 {PAGE}。把它放到本程序（exe）同目录即可。",
                        status=500, mimetype="text/plain; charset=utf-8")


@app.route("/health")
def health():
    return {"ok": True}


# ---------- 邮件：状态 / 登录 / 发送 ----------
@app.route("/mail/status")
def mail_status():
    return {
        "configured": bool(GRAPH_CFG["client_id"]),
        "signedIn": bool(GRAPH_TOKEN["access"] or GRAPH_TOKEN["refresh"]),
        "user": GRAPH_TOKEN["user"] or "",
        "canDraft": bool(GRAPH_CFG["allow_draft"]),
    }


# ---------- 登录方式 A：浏览器直接登录（授权码 + PKCE）—— 不用抄任何代码 ----------
@app.route("/mail/authurl")
def mail_authurl():
    if not GRAPH_CFG["client_id"]:
        return {"ok": False, "msg": "还没配置 client_id（把 graph_config.json 放到程序同目录）。"}
    verifier = secrets.token_urlsafe(64)
    challenge = base64.urlsafe_b64encode(
        hashlib.sha256(verifier.encode("ascii")).digest()).decode("ascii").rstrip("=")
    state = secrets.token_urlsafe(16)
    OAUTH["verifier"], OAUTH["state"] = verifier, state
    q = urlencode({
        "client_id": GRAPH_CFG["client_id"],
        "response_type": "code",
        "redirect_uri": REDIRECT_URI,
        "response_mode": "query",
        "scope": _graph_scopes(),
        "state": state,
        "code_challenge": challenge,
        "code_challenge_method": "S256",
        "prompt": "select_account",
    })
    return {"ok": True,
            "url": f"{_authority()}/oauth2/v2.0/authorize?{q}",
            "redirect": REDIRECT_URI}


def _cb_page(ok, msg):
    color = "#1f7a4d" if ok else "#b3352f"
    title = "登录成功" if ok else "登录失败"
    body = (f"已登录：{msg}<br>这个标签页可以关掉了，回到工具页面继续。" if ok
            else f"{msg}<br><br>回到工具页面重试。")
    html = f"""<!DOCTYPE html><html lang="zh"><head><meta charset="UTF-8">
<title>{title}</title><style>
body{{font-family:-apple-system,"Segoe UI",system-ui,"PingFang SC","Microsoft YaHei",sans-serif;
background:#F3F6FA;display:grid;place-items:center;height:100vh;margin:0;}}
.card{{background:#fff;border:1px solid #E2E7ED;border-radius:14px;padding:34px 40px;max-width:520px;
text-align:center;box-shadow:0 2px 10px rgba(20,30,50,.06);}}
h1{{color:{color};margin:0 0 12px;font-size:22px;}}
p{{color:#4a5561;font-size:14px;line-height:1.7;margin:0;word-break:break-word;}}
</style></head><body><div class="card"><h1>{title}</h1><p>{body}</p></div>
<script>if({str(ok).lower()}) setTimeout(()=>window.close(),2500);</script>
</body></html>"""
    return Response(html, mimetype="text/html; charset=utf-8")


@app.route("/mail/callback")
def mail_callback():
    if request.args.get("error"):
        return _cb_page(False, f"{request.args.get('error')}: "
                               f"{(request.args.get('error_description') or '')[:400]}")
    code = request.args.get("code")
    state = request.args.get("state")
    if not code or not OAUTH["state"] or state != OAUTH["state"]:
        return _cb_page(False, "state 校验没过（可能是过期的登录链接）。回到工具页面重新点登录。")
    try:
        r = requests.post(f"{_authority()}/oauth2/v2.0/token", data={
            "client_id": GRAPH_CFG["client_id"],
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": REDIRECT_URI,
            "code_verifier": OAUTH["verifier"],
            "scope": _graph_scopes(),
        }, timeout=30)
        j = r.json()
    except Exception as e:
        return _cb_page(False, f"换 token 失败：{e}")
    OAUTH["verifier"] = OAUTH["state"] = None
    if "access_token" not in j:
        return _cb_page(False, j.get("error_description") or str(j)[:400])
    _store_token(j)
    return _cb_page(True, GRAPH_TOKEN["user"] or "")


# ---------- 登录方式 B：设备码（备用；Azure 没登记重定向 URI 时用）----------
@app.route("/mail/signin-start", methods=["POST"])
def mail_signin_start():
    if not GRAPH_CFG["client_id"]:
        return {"ok": False, "msg": "还没配置 client_id（把 graph_config.json 放到程序同目录）。"}
    try:
        r = requests.post(f"{_authority()}/oauth2/v2.0/devicecode",
                          data={"client_id": GRAPH_CFG["client_id"], "scope": _graph_scopes()},
                          timeout=30)
        j = r.json()
    except Exception as e:
        return {"ok": False, "msg": f"连不上 Microsoft 登录服务：{e}"}
    if "device_code" not in j:
        return {"ok": False, "msg": j.get("error_description") or str(j)[:300]}
    GRAPH_FLOW["device_code"] = j["device_code"]
    GRAPH_FLOW["interval"] = int(j.get("interval", 5))
    GRAPH_FLOW["expires"] = time.time() + int(j.get("expires_in", 900))
    return {"ok": True,
            "userCode": j.get("user_code"),
            "verificationUri": j.get("verification_uri") or "https://microsoft.com/devicelogin",
            "interval": GRAPH_FLOW["interval"]}


@app.route("/mail/signin-poll", methods=["POST"])
def mail_signin_poll():
    if not GRAPH_FLOW["device_code"]:
        return {"ok": False, "state": "none", "msg": "还没开始登录。"}
    if time.time() > GRAPH_FLOW["expires"]:
        GRAPH_FLOW["device_code"] = None
        return {"ok": False, "state": "expired", "msg": "登录码过期了，重新点一次登录。"}
    try:
        r = requests.post(f"{_authority()}/oauth2/v2.0/token", data={
            "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
            "client_id": GRAPH_CFG["client_id"],
            "device_code": GRAPH_FLOW["device_code"],
        }, timeout=30)
        j = r.json()
    except Exception as e:
        return {"ok": False, "state": "error", "msg": str(e)}
    if "access_token" in j:
        GRAPH_FLOW["device_code"] = None
        _store_token(j)
        return {"ok": True, "state": "done", "user": GRAPH_TOKEN["user"]}
    err = j.get("error", "")
    if err in ("authorization_pending", "slow_down"):
        return {"ok": True, "state": "pending"}
    GRAPH_FLOW["device_code"] = None
    return {"ok": False, "state": "error",
            "msg": j.get("error_description") or err or "登录失败"}


@app.route("/mail/signout", methods=["POST"])
def mail_signout():
    GRAPH_TOKEN.update({"access": None, "refresh": None, "exp": 0.0, "user": ""})
    GRAPH_FLOW["device_code"] = None
    return {"ok": True}


@app.route("/mail/send-one", methods=["POST"])
def mail_send_one():
    """发一封（或建一封草稿）。前端循环调用，好显示进度、也好定位是哪封失败。"""
    d = request.get_json(force=True, silent=True) or {}
    mode = (d.get("mode") or "draft").lower()
    to = (d.get("to") or "").strip()
    if not to:
        return {"ok": False, "msg": "没有收件人"}
    if mode == "draft" and not GRAPH_CFG["allow_draft"]:
        return {"ok": False, "msg": "草稿功能没开（需要 Mail.ReadWrite 权限）"}

    file_b64 = d.get("fileB64") or ""
    file_name = d.get("fileName") or "attachment.xlsx"
    if len(file_b64) > 3_000_000:                      # Graph 内联附件上限 ~3MB
        return {"ok": False, "msg": f"{file_name} 超过 3MB，Graph 内联附件放不下"}

    msg = {
        "subject": d.get("subject") or "",
        "body": {"contentType": "Text", "content": d.get("body") or ""},
        "toRecipients": _recips(to),
        "attachments": [{
            "@odata.type": "#microsoft.graph.fileAttachment",
            "name": file_name,
            "contentType": XLSX_MIME,
            "contentBytes": file_b64,
        }],
    }
    cc = (d.get("cc") or "").strip()
    if cc:
        msg["ccRecipients"] = _recips(cc)

    try:
        tok = graph_token()
    except Exception as e:
        return {"ok": False, "needLogin": True, "msg": str(e)}

    h = {"Authorization": "Bearer " + tok, "Content-Type": "application/json"}
    try:
        if mode == "send":
            r = requests.post(f"{GRAPH_API}/me/sendMail", headers=h,
                              json={"message": msg, "saveToSentItems": True}, timeout=120)
            ok = r.status_code in (200, 202)
        else:
            r = requests.post(f"{GRAPH_API}/me/messages", headers=h, json=msg, timeout=120)
            ok = r.status_code in (200, 201)
        if ok:
            return {"ok": True}
        return {"ok": False, "msg": f"Graph {r.status_code}: {r.text[:300]}"}
    except Exception as e:
        return {"ok": False, "msg": f"{type(e).__name__}: {e}"}


# ===================== 启动 =====================
def _port_busy(port):
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(("127.0.0.1", port)) == 0


def _open_browser(port):
    webbrowser.open(f"http://127.0.0.1:{port}/")


if __name__ == "__main__":
    ok_graph = load_graph_cfg()
    port = CFG["port"]

    if _port_busy(port):
        print(f"端口 {port} 已经被占用了 —— 是不是已经开着一个了？")
        print("（这个端口不能换：Azure 里登记的登录回调地址是写死这个端口的。）")
        print("关掉那个再开，或者重启电脑。")
        input("按回车关闭…")
        sys.exit(1)

    print("=" * 58)
    print("  SpeedX 包裹清单拆分器 + 按车队发邮件")
    print("=" * 58)
    print(f"  打开： http://127.0.0.1:{port}/")
    print(f"  页面： {PAGE}  ← {asset_origin(PAGE)}")
    if ok_graph:
        cid = GRAPH_CFG["client_id"]
        print(f"  邮件： Graph 已配置 (client_id …{cid[-6:]})")
        print(f"  登录回调： {REDIRECT_URI}")
        print("           ↑ 这一条必须原样登记在 Azure → Authentication 里")
    else:
        print("  邮件： 未配置 client_id → 拆分照常用，但『发送』是灰的。")
        print("         把 graph_config.json 放到本程序同目录即可启用。")
    print("=" * 58)
    print("  关掉这个黑框 = 退出程序。")
    print()

    threading.Timer(1.2, _open_browser, args=(port,)).start()
    app.run(host="127.0.0.1", port=port, debug=False)
